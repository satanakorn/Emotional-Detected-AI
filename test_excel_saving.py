#!/usr/bin/env python3

import os
import time
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

class ExcelSavingTester:
    def __init__(self):
        self.excel_file = "test_emotion_data.xlsx"
        self._wb = None
        
    def create_test_excel(self):
        """สร้างไฟล์ Excel ทดสอบ"""
        try:
            print(f"🔍 Creating test Excel file: {self.excel_file}")
            
            self._wb = Workbook()
            ws = self._wb.active
            ws.title = "ข้อมูลอารมณ์"
            
            headers = [
                "วันที่", "เวลา", "อารมณ์", "ความมั่นใจ (%)", 
                "ระดับความพึงพอใจ", "คะแนน", "วิธีการกล้อง"
            ]
            
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[chr(64 + col)].width = 15
            
            self._wb.save(self.excel_file)
            print(f"✅ Test Excel file created: {self.excel_file}")
            print(f"📁 File path: {os.path.abspath(self.excel_file)}")
            return True
            
        except Exception as e:
            print(f"❌ Error creating Excel file: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def add_test_data(self, emotion, confidence, satisfaction_level, satisfaction_text, camera_method):
        """เพิ่มข้อมูลทดสอบลง Excel"""
        try:
            if not os.path.exists(self.excel_file):
                print("❌ Excel file not found, creating new one...")
                if not self.create_test_excel():
                    return False
            
            # โหลดไฟล์ Excel
            if self._wb is None:
                self._wb = openpyxl.load_workbook(self.excel_file)
            
            ws = self._wb.active
            
            # หาแถวถัดไป
            next_row = ws.max_row + 1
            
            # เตรียมข้อมูล
            now = datetime.now()
            data = [
                now.strftime("%Y-%m-%d"),
                now.strftime("%H:%M:%S"),
                emotion,
                f"{confidence:.2f}",
                satisfaction_level,
                satisfaction_text,
                camera_method
            ]
            
            # บันทึกข้อมูล
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=next_row, column=col)
                cell.value = value
                cell.alignment = Alignment(horizontal='center')
            
            # บันทึกไฟล์
            self._wb.save(self.excel_file)
            
            print(f"✅ Data added to Excel row {next_row}: {emotion} ({confidence:.1f}%)")
            return True
            
        except Exception as e:
            print(f"❌ Error adding data to Excel: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def test_excel_operations(self):
        """ทดสอบการทำงานของ Excel"""
        print("🧪 Testing Excel Operations")
        print("=" * 40)
        
        # ทดสอบการสร้างไฟล์
        if not self.create_test_excel():
            return False
        
        # ทดสอบการเพิ่มข้อมูล
        test_data = [
            ("Happy", 85.5, 5, "*****", "test_camera"),
            ("Neutral", 60.0, 3, "***", "test_camera"),
            ("Sad", 35.2, 2, "**", "test_camera"),
            ("Angry", 15.8, 1, "*", "test_camera")
        ]
        
        print("\n📝 Adding test data...")
        for emotion, confidence, level, text, camera in test_data:
            success = self.add_test_data(emotion, confidence, level, text, camera)
            if not success:
                print(f"❌ Failed to add data: {emotion}")
                return False
            time.sleep(0.5)  # รอเล็กน้อย
        
        # ตรวจสอบไฟล์
        if os.path.exists(self.excel_file):
            file_size = os.path.getsize(self.excel_file)
            print(f"\n✅ Excel file created successfully!")
            print(f"📁 File: {os.path.abspath(self.excel_file)}")
            print(f"📊 File size: {file_size} bytes")
            print(f"📝 Data rows: {len(test_data) + 1} (including header)")
            
            # อ่านข้อมูลกลับมาเพื่อตรวจสอบ
            try:
                wb = openpyxl.load_workbook(self.excel_file)
                ws = wb.active
                print(f"📋 Sheet name: {ws.title}")
                print(f"📊 Total rows: {ws.max_row}")
                print(f"📊 Total columns: {ws.max_column}")
                
                # แสดงข้อมูลบางส่วน
                print("\n📋 Sample data:")
                for row in range(1, min(4, ws.max_row + 1)):
                    row_data = []
                    for col in range(1, ws.max_column + 1):
                        cell_value = ws.cell(row=row, column=col).value
                        row_data.append(str(cell_value))
                    print(f"   Row {row}: {' | '.join(row_data)}")
                
                return True
                
            except Exception as e:
                print(f"❌ Error reading Excel file: {e}")
                return False
        else:
            print("❌ Excel file not found after creation")
            return False
    
    def cleanup(self):
        """ลบไฟล์ทดสอบ"""
        try:
            if os.path.exists(self.excel_file):
                os.remove(self.excel_file)
                print(f"🗑️ Test file removed: {self.excel_file}")
        except Exception as e:
            print(f"⚠️ Error removing test file: {e}")

def main():
    print("🧪 Excel Saving Test Tool")
    print("🎯 Test Excel file creation and data saving")
    print("=" * 50)
    
    tester = ExcelSavingTester()
    
    try:
        success = tester.test_excel_operations()
        
        if success:
            print("\n✅ Excel operations test completed successfully!")
            print("💡 The Excel saving functionality is working correctly.")
        else:
            print("\n❌ Excel operations test failed!")
            print("💡 Check the error messages above for details.")
        
        # ถามว่าต้องการลบไฟล์ทดสอบหรือไม่
        try:
            keep_file = input("\n🗑️ Keep test file? (y/n): ").strip().lower()
            if keep_file != 'y':
                tester.cleanup()
        except:
            tester.cleanup()
            
    except KeyboardInterrupt:
        print("\n⚠️ Test interrupted")
        tester.cleanup()
    except Exception as e:
        print(f"\n❌ Unexpected error: {e}")
        tester.cleanup()

if __name__ == "__main__":
    main()
