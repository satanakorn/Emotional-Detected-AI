#!/usr/bin/env python3

import cv2
import numpy as np
import time
import os
import sys
from datetime import datetime
import subprocess
import threading
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from collections import deque
import queue

# ค่าคงที่สำหรับการปรับแต่งประสิทธิภาพ
FRAME_SKIP = 1  # ไม่ข้ามเฟรม (real-time detection)
EMOTION_CACHE_SIZE = 3  # ลดแคชเพื่อความเร็ว
EXCEL_SAVE_INTERVAL = 5  # บันทึก Excel ทุก 5 วินาที
MAX_QUEUE_SIZE = 100  # ขนาดสูงสุดของคิวสำหรับการบันทึกข้อมูล
DETECTION_INTERVAL = 0.05  # ตรวจจับทุก 50ms (20 FPS)
DATA_SAVE_INTERVAL = 5.0  # บันทึกข้อมูลทุก 5 วินาที

try:
    from picamera2 import Picamera2
    from libcamera import controls
    PICAMERA2_AVAILABLE = True
    print("✅ PiCamera2 library loaded successfully")
except ImportError:
    PICAMERA2_AVAILABLE = False
    print("⚠️ PiCamera2 not available, falling back to OpenCV")

try:
    from deepface import DeepFace
    DEEPFACE_AVAILABLE = True
    print("✅ DeepFace library loaded successfully")
except ImportError:
    DEEPFACE_AVAILABLE = False
    print("⚠️ DeepFace not installed. Using simple face detection only.")
    print("Install with: pip install deepface tensorflow")

class RaspberryPi4CameraDetector:
    def __init__(self):
        self.cap = None
        self.picam2 = None
        self.camera_method = None
        self.face_cascade = None
        self.emotion_history = deque(maxlen=100)  # ใช้ deque แทน list
        self.is_running = False
        self.frame_buffer = None
        self.buffer_lock = threading.Lock()
        self.color_mode = "color"
        self.auto_exposure = True
        self.brightness = 0.0
        self.contrast = 1.0
        self.camera_type = None
        self.excel_file = "emotion_data.xlsx"
        
        # เพิ่มตัวแปรสำหรับการปรับแต่งประสิทธิภาพ
        self.frame_count = 0
        self.last_emotion_time = 0
        self.last_data_save_time = time.time()  # เวลาบันทึกข้อมูลล่าสุด
        self.emotion_cache = {}  # แคชผลการตรวจจับอารมณ์
        self.data_queue = queue.Queue(maxsize=MAX_QUEUE_SIZE)
        self.excel_save_counter = 0
        self.current_emotion_data = None  # ข้อมูลอารมณ์ปัจจุบัน
        
        # เพิ่มตัวแปรสำหรับ manual emotion input
        self.manual_emotion = None
        self.manual_confidence = 0.0
        self.manual_mode = False
        self.last_manual_time = 0
        
        # เพิ่มตัวแปรสำหรับการบันทึก Excel แบบ time-based
        self.last_excel_save_time = time.time()
        
        # เพิ่มตัวแปรสำหรับการแสดงผลเต็มจอ
        self.fullscreen_mode = True
        
        # เริ่มเธรดสำหรับการบันทึกข้อมูล
        self.save_thread = threading.Thread(target=self._save_data_worker, daemon=True)
        self.save_thread.start()
        
        self.load_face_cascade()
        self.initialize_excel()

    def _save_data_worker(self):
        """เธรดสำหรับการบันทึกข้อมูลลง Excel"""
        while self.is_running:
            try:
                data = self.data_queue.get(timeout=1)
                if data:
                    self._save_to_excel_internal(*data)
                    self.excel_save_counter += 1
                    
                    # บันทึกไฟล์ Excel ทุกๆ 5 วินาที
                    current_time = time.time()
                    if current_time - self.last_excel_save_time >= EXCEL_SAVE_INTERVAL:
                        self._save_excel_file()
                        self.last_excel_save_time = current_time
                        print(f"💾 Excel saved - {self.excel_save_counter} records")
            except queue.Empty:
                continue
            except Exception as e:
                print(f"❌ เกิดข้อผิดพลาดในการบันทึกข้อมูล: {e}")

    def _save_excel_file(self):
        """บันทึกไฟล์ Excel"""
        try:
            if hasattr(self, '_wb'):
                self._wb.save(self.excel_file)
        except Exception as e:
            print(f"❌ เกิดข้อผิดพลาดในการบันทึกไฟล์ Excel: {e}")

    def initialize_excel(self):
        """สร้างไฟล์ Excel ใหม่ถ้ายังไม่มี"""
        try:
            print(f"🔍 Checking Excel file: {self.excel_file}")
            
            if not os.path.exists(self.excel_file):
                print("📝 Creating new Excel file...")
                self._wb = Workbook()
                ws = self._wb.active
                ws.title = "ข้อมูลอารมณ์"
                
                headers = [
                    "วันที่", "เวลา", "อารมณ์", "ความมั่นใจ (%)", 
                    "ระดับความพึงพอใจ", "คะแนน", "วิธีการกล้อง"
                ]
                
                header_font = Font(bold=True, size=12)
                header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                
                for col in range(1, len(headers) + 1):
                    ws.column_dimensions[chr(64 + col)].width = 15
                
                self._wb.save(self.excel_file)
                print(f"✅ สร้างไฟล์ Excel ใหม่: {self.excel_file}")
                print(f"📁 File path: {os.path.abspath(self.excel_file)}")
            else:
                print(f"✅ Excel file already exists: {self.excel_file}")
                self._wb = openpyxl.load_workbook(self.excel_file)
                
        except Exception as e:
            print(f"❌ เกิดข้อผิดพลาดในการสร้างไฟล์ Excel: {e}")
            import traceback
            traceback.print_exc()

    def save_to_excel(self, emotion, confidence, satisfaction_level, satisfaction_text):
        """เพิ่มข้อมูลลงในคิวสำหรับการบันทึก (ทุก 5 วินาที)"""
        try:
            current_time = time.time()
            
            # เก็บข้อมูลปัจจุบันไว้
            self.current_emotion_data = {
                'emotion': emotion,
                'confidence': confidence,
                'satisfaction_level': satisfaction_level,
                'satisfaction_text': satisfaction_text,
                'timestamp': current_time
            }
            
            # บันทึกข้อมูลทุก 5 วินาที
            if current_time - self.last_data_save_time >= DATA_SAVE_INTERVAL:
                now = datetime.now()
                data = (
                    now.strftime("%Y-%m-%d"),
                    now.strftime("%H:%M:%S"),
                    emotion,
                    f"{confidence:.2f}",
                    satisfaction_level,
                    satisfaction_text,
                    self.camera_method
                )
                
                # บันทึกข้อมูลลง Excel โดยตรง
                self._save_to_excel_internal(*data)
                self._save_excel_file()
                
                self.last_data_save_time = current_time
                print(f"💾 Data saved to Excel: {emotion} ({confidence:.1f}%)")
                
        except Exception as e:
            print(f"❌ Error saving to Excel: {e}")

    def _save_to_excel_internal(self, date, time, emotion, confidence, satisfaction_level, satisfaction_text, camera_method):
        """บันทึกข้อมูลลงในไฟล์ Excel"""
        try:
            # ตรวจสอบว่ามีไฟล์ Excel หรือไม่
            if not os.path.exists(self.excel_file):
                print("📝 Creating new Excel file...")
                self.initialize_excel()
            
            # โหลดไฟล์ Excel
            if not hasattr(self, '_wb') or self._wb is None:
                self._wb = openpyxl.load_workbook(self.excel_file)
            
            ws = self._wb.active
            
            # หาแถวถัดไป
            next_row = ws.max_row + 1
            data = [date, time, emotion, confidence, satisfaction_level, satisfaction_text, camera_method]
            
            # บันทึกข้อมูล
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=next_row, column=col)
                cell.value = value
                cell.alignment = Alignment(horizontal='center')
            
            print(f"✅ Data added to Excel row {next_row}: {emotion} ({confidence}%)")
            
        except Exception as e:
            print(f"❌ เกิดข้อผิดพลาดในการบันทึกข้อมูลลง Excel: {e}")
            import traceback
            traceback.print_exc()

    def load_face_cascade(self):
        """แก้ปัญหา cv2.data ใน Raspberry Pi 4 และโหลด cascade หลายแบบ"""
        try:
            # ลองหา cascade files ในตำแหน่งต่างๆ   
            possible_paths = [
                '/usr/share/opencv4/haarcascades/haarcascade_frontalface_default.xml',
                '/usr/local/share/opencv4/haarcascades/haarcascade_frontalface_default.xml',
                '/home/pi/.local/lib/python3.*/site-packages/cv2/data/haarcascade_frontalface_default.xml'
            ]
            
            # ลองใช้ cv2.data ก่อน
            try:
                cascade_path = cv2.data.haarcascades + 'haarcascade_frontalface_default.xml'
                if os.path.exists(cascade_path):
                    self.face_cascade = cv2.CascadeClassifier(cascade_path)
                    print("✅ Face cascade loaded from cv2.data")
                    return
            except AttributeError:
                pass
            
            # ถ้าไม่ได้ ให้ลองจากตำแหน่งอื่น
            for path in possible_paths:
                if os.path.exists(path):
                    self.face_cascade = cv2.CascadeClassifier(path)
                    print(f"✅ Face cascade loaded from {path}")
                    return
            
            # ดาวน์โหลดถ้าไม่เจอ
            print("📥 Downloading Haar cascade file...")
            import urllib.request
            url = "https://raw.githubusercontent.com/opencv/opencv/master/data/haarcascades/haarcascade_frontalface_default.xml"
            cascade_path = "haarcascade_frontalface_default.xml"
            urllib.request.urlretrieve(url, cascade_path)
            self.face_cascade = cv2.CascadeClassifier(cascade_path)
            print("✅ Face cascade downloaded and loaded")
            
        except Exception as e:
            print(f"❌ Error loading face cascade: {e}")
            self.face_cascade = None
    
    def enhance_frame_for_detection(self, frame):
        """ปรับปรุงเฟรมเพื่อการตรวจจับใบหน้าที่ดีขึ้น"""
        try:
            # ปรับความสว่างและคอนทราสต์
            enhanced = cv2.convertScaleAbs(frame, alpha=1.2, beta=10)
            
            # ลด noise
            enhanced = cv2.bilateralFilter(enhanced, 9, 75, 75)
            
            # ปรับ histogram
            if len(enhanced.shape) == 3:
                # แปลงเป็น LAB color space
                lab = cv2.cvtColor(enhanced, cv2.COLOR_BGR2LAB)
                l, a, b = cv2.split(lab)
                
                # ปรับ L channel ด้วย CLAHE
                clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
                l = clahe.apply(l)
                
                # รวมกลับ
                enhanced = cv2.merge([l, a, b])
                enhanced = cv2.cvtColor(enhanced, cv2.COLOR_LAB2BGR)
            
            return enhanced
        except Exception as e:
            print(f"Frame enhancement error: {e}")
            return frame
    
    def check_camera_hardware(self):
        """ตรวจสอบฮาร์ดแวร์กล้องใน Raspberry Pi 4"""
        print("🔍 Enhanced camera status check...")
        print("=" * 50)
        
        # ตรวจสอบ libcamera
        try:
            result = subprocess.run(['libcamera-hello', '--list-cameras'], 
                                  capture_output=True, text=True, timeout=10)
            if result.returncode == 0:
                print("✅ libcamera tools available")
                print("📷 Available cameras:")
                print(result.stdout)
            else:
                print("⚠️ libcamera available but no cameras detected")
        except (subprocess.TimeoutExpired, FileNotFoundError):
            print("❌ libcamera tools not found")
        
        # ตรวจสอบ video devices
        print("\n📹 Video devices:")
        os.system("ls -la /dev/video* 2>/dev/null || echo '⚠️ No video devices found'")
        
        # ตรวจสอบ vcgencmd
        print("\n⚙️ Camera configuration:")
        os.system("vcgencmd get_camera 2>/dev/null || echo '⚠️ Cannot get camera status'")
        
        # ตรวจสอบ config.txt
        print("\n📝 Config.txt camera settings:")
        os.system("grep -E '(camera|start_x|gpu_mem)' /boot/config.txt 2>/dev/null || echo '⚠️ No camera settings found'")
        
        print("=" * 50)
    
    def setup_picamera2(self):
        """ตั้งค่า PiCamera2 สำหรับ Raspberry Pi 4 พร้อมจัดการสี"""
        try:
            if not PICAMERA2_AVAILABLE:
                return False
            
            print("🔄 Trying PiCamera2 method...")
            
            self.picam2 = Picamera2()
            
            # กำหนดค่า preview config สำหรับสีที่ถูกต้อง
            if self.color_mode == "grayscale":
                config = self.picam2.create_preview_configuration(
                    main={"size": (640, 480), "format": "YUV420"},
                    controls={"FrameRate": 30}
                )
            else:
                config = self.picam2.create_preview_configuration(
                    main={"size": (640, 480), "format": "RGB888"},
                    controls={"FrameRate": 30}
                )
            
            self.picam2.configure(config)
            
            # ตั้งค่า controls สำหรับการปรับแสงและสี
            control_settings = {
                "AeEnable": self.auto_exposure,
                "AwbEnable": True,  # Auto White Balance
                "AwbMode": controls.AwbModeEnum.Auto,
                "Brightness": self.brightness,
                "Contrast": self.contrast,
                "Saturation": 1.0,
                "Sharpness": 1.0
            }
            
            self.picam2.set_controls(control_settings)
            
            self.picam2.start()
            time.sleep(2)  # รอให้กล้องเริ่มต้นและปรับแสง
            
            # ทดสอบการจับภาพ
            frame = self.picam2.capture_array()
            if frame is not None and frame.size > 0:
                print(f"✅ PiCamera2 method successful - Format: {frame.shape}")
                print(f"📷 Color mode: {self.color_mode}")
                self.camera_method = "picamera2"
                self.start_frame_capture_thread()
                return True
            else:
                self.picam2.stop()
                return False
                
        except Exception as e:
            print(f"   PiCamera2 error: {e}")
            if self.picam2:
                try:
                    self.picam2.stop()
                except:
                    pass
            return False
    
    def start_frame_capture_thread(self):
        """เริ่มเธรดสำหรับจับภาพอย่างต่อเนื่อง"""
        def capture_frames():
            while self.is_running:
                try:
                    if self.picam2:
                        frame = self.picam2.capture_array()
                        with self.buffer_lock:
                            self.frame_buffer = frame.copy()
                    time.sleep(0.03)  # ~30 FPS
                except Exception as e:
                    print(f"Frame capture error: {e}")
                    break
        
        self.is_running = True
        self.capture_thread = threading.Thread(target=capture_frames, daemon=True)
        self.capture_thread.start()
    
    def setup_opencv_camera(self):
        """ตั้งค่ากล้องด้วย OpenCV สำหรับ Raspberry Pi 4 พร้อมจัดการสี"""
        try:
            print("🔄 Trying simple OpenCV camera access...")
            
            # ลองกล้อง USB แบบธรรมดาก่อน (ไม่ใช้ V4L2 หรือ GStreamer)
            for i in range(5):  # ลอง 5 devices
                print(f"   Testing camera index {i}...")
                self.cap = cv2.VideoCapture(i)
                
                if self.cap.isOpened():
                    # ตั้งค่าความละเอียดและ FPS
                    self.cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
                    self.cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
                    self.cap.set(cv2.CAP_PROP_FPS, 30)
                    self.cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)
                    
                    ret, frame = self.cap.read()
                    if ret and frame is not None:
                        print(f"✅ USB camera found at index {i} - Shape: {frame.shape}")
                        self.camera_method = f"usb_{i}"
                        return True
                    self.cap.release()
                else:
                    print(f"   Camera {i} not available")
            
            # ถ้าไม่ได้ ลอง GStreamer pipeline
            print("🔄 Trying GStreamer pipeline...")
            try:
                if self.color_mode == "grayscale":
                    gst_pipeline = (
                        "libcamerasrc ! "
                        "video/x-raw,width=640,height=480,framerate=30/1,format=GRAY8 ! "
                        "videoconvert ! appsink"
                    )
                else:
                    gst_pipeline = (
                        "libcamerasrc ! "
                        "video/x-raw,width=640,height=480,framerate=30/1 ! "
                        "videoconvert ! appsink"
                    )
                
                self.cap = cv2.VideoCapture(gst_pipeline, cv2.CAP_GSTREAMER)
                
                if self.cap.isOpened():
                    ret, frame = self.cap.read()
                    if ret and frame is not None:
                        print(f"✅ GStreamer pipeline successful - Shape: {frame.shape}")
                        self.camera_method = "gstreamer"
                        return True
                    self.cap.release()
            except Exception as e:
                print(f"   GStreamer error: {e}")
            
            # ถ้าไม่ได้ ลอง V4L2
            print("🔄 Trying V4L2 direct access...")
            try:
                # เปิดใช้งาน bcm2835-v4l2 module
                os.system("sudo modprobe bcm2835-v4l2 2>/dev/null")
                time.sleep(1)
                
                self.cap = cv2.VideoCapture(0, cv2.CAP_V4L2)
                
                if self.cap.isOpened():
                    # ตั้งค่าความละเอียดและ FPS
                    self.cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
                    self.cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
                    self.cap.set(cv2.CAP_PROP_FPS, 30)
                    self.cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)
                    
                    ret, frame = self.cap.read()
                    if ret and frame is not None:
                        print(f"✅ V4L2 method successful - Shape: {frame.shape}")
                        self.camera_method = "v4l2"
                        return True
                    self.cap.release()
            except Exception as e:
                print(f"   V4L2 error: {e}")
            
            print("❌ No working camera found with OpenCV")
            return False
            
        except Exception as e:
            print(f"   OpenCV camera error: {e}")
            if self.cap:
                self.cap.release()
            return False
    
    def setup_laptop_camera(self):
        """ตั้งค่ากล้องโน๊ตบุ๊ค"""
        try:
            print("🔄 กำลังตั้งค่ากล้องเว็บแคม...")
            self.cap = cv2.VideoCapture(0)  # ใช้กล้องแรกที่เจอ
            
            if self.cap.isOpened():
                # ตั้งค่าความละเอียดและ FPS
                self.cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
                self.cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
                self.cap.set(cv2.CAP_PROP_FPS, 30)
                
                ret, frame = self.cap.read()
                if ret and frame is not None:
                    print(f"✅ ตั้งค่ากล้องเว็บแคมสำเร็จ - ขนาดภาพ: {frame.shape}")
                    self.camera_method = "laptop_webcam"
                    return True
                self.cap.release()
            
            return False
            
        except Exception as e:
            print(f"❌ เกิดข้อผิดพลาดในการตั้งค่ากล้องเว็บแคม: {e}")
            if self.cap:
                self.cap.release()
            return False
    
    def detect_usb_camera(self):
        """ตรวจสอบว่ามีกล้อง USB หรือไม่"""
        try:
            print("🔍 Checking for USB cameras...")
            
            # ลองเปิดกล้องแบบธรรมดาก่อน
            for i in range(5):
                cap = cv2.VideoCapture(i)
                if cap.isOpened():
                    ret, frame = cap.read()
                    if ret and frame is not None:
                        print(f"✅ Working camera found at index {i}")
                        cap.release()
                        return True
                    cap.release()
            
            # ตรวจสอบ video devices
            import glob
            video_devices = glob.glob('/dev/video*')
            if video_devices:
                print(f"📹 Found video devices: {video_devices}")
                
                # ลองเปิดกล้องแต่ละตัว
                for device in video_devices:
                    try:
                        device_num = int(device.split('video')[1])
                        cap = cv2.VideoCapture(device_num)
                        if cap.isOpened():
                            ret, frame = cap.read()
                            if ret and frame is not None:
                                print(f"✅ Working camera found at {device}")
                                cap.release()
                                return True
                            cap.release()
                    except:
                        continue
            
            # ตรวจสอบ USB devices
            import subprocess
            try:
                result = subprocess.run(['lsusb'], capture_output=True, text=True, timeout=5)
                if 'camera' in result.stdout.lower() or 'webcam' in result.stdout.lower():
                    print("📷 USB camera detected in lsusb")
                    return True
            except:
                pass
            
            print("⚠️ No USB cameras detected")
            return False
            
        except Exception as e:
            print(f"❌ Error detecting USB camera: {e}")
            return False
    
    def set_manual_emotion(self, emotion, confidence):
        """ตั้งค่าอารมณ์แบบ manual"""
        self.manual_emotion = emotion
        self.manual_confidence = confidence
        self.last_manual_time = time.time()
        print(f"🎯 Manual emotion set: {emotion} ({confidence:.1f}%)")
    
    def get_emotion_assessment(self, confidence):
        """ประเมินอารมณ์ตามเกณฑ์ใหม่"""
        if confidence >= 80:
            return "Happy", "*****", 5
        elif confidence >= 50:
            return "Neutral", "***", 3
        elif confidence >= 30:
            return "Sad", "**", 2
        else:
            return "Angry", "*", 1
    
    def setup_camera(self):
        """ตั้งค่ากล้องตามประเภทที่เลือก"""
        print("🔍 Camera setup...")
        
        if self.camera_type == "laptop":
            return self.setup_laptop_camera()
        elif self.camera_type == "pi":
            # สำหรับกล้อง USB ให้ลอง OpenCV ก่อน
            print("🔌 Detecting camera type...")
            
            # ตรวจสอบว่ามีกล้อง USB หรือไม่
            usb_camera_found = self.detect_usb_camera()
            
            if usb_camera_found:
                print("📷 USB camera detected, using OpenCV method")
                if self.setup_opencv_camera():
                    return True
            
            # ถ้าไม่มีกล้อง USB หรือไม่ทำงาน ลอง PiCamera2 (สำหรับ CSI camera)
            print("📷 Trying PiCamera2 for CSI camera...")
            if self.setup_picamera2():
                return True
            
            # ถ้า PiCamera2 ไม่ได้ ลอง OpenCV อีกครั้ง
            print("🔄 Fallback to OpenCV...")
            if self.setup_opencv_camera():
                return True
        
        print("❌ No camera found. Troubleshooting steps:")
        print("   1. For USB cameras: Check USB connection and permissions")
        print("   2. For Pi Camera Module: Check CSI cable and enable camera")
        print("   3. Enable camera: sudo raspi-config")
        print("   4. Add to /boot/config.txt: camera_auto_detect=1")
        print("   5. Increase GPU memory: gpu_mem=128")
        print("   6. Reboot system")
        print("   7. Run diagnostic: python3 diagnose_camera.py")
        return False
    
    def get_frame(self):
        """รับเฟรมจากกล้องและจัดการสี"""
        frame = None
        
        if self.camera_method == "picamera2":
            with self.buffer_lock:
                if self.frame_buffer is not None:
                    frame = self.frame_buffer.copy()
        elif self.cap and self.cap.isOpened():
            ret, frame = self.cap.read()
            if not ret:
                frame = None
        
        if frame is None:
            return None
        
        # จัดการการแปลงสีตามโหมดที่เลือก
        if self.color_mode == "grayscale":
            if len(frame.shape) == 3:
                if self.camera_method == "picamera2":
                    # PiCamera2 RGB to Grayscale
                    frame = cv2.cvtColor(frame, cv2.COLOR_RGB2GRAY)
                else:
                    # OpenCV BGR to Grayscale
                    frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                # แปลงกลับเป็น 3 channels สำหรับการแสดงผล
                frame = cv2.cvtColor(frame, cv2.COLOR_GRAY2BGR)
        
        elif self.color_mode == "color":
            # ตรวจสอบให้แน่ใจว่าเป็น BGR สำหรับ OpenCV
            if self.camera_method == "picamera2" and len(frame.shape) == 3:
                # แปลงจาก RGB เป็น BGR
                frame = cv2.cvtColor(frame, cv2.COLOR_RGB2BGR)
        
        return frame
    
    def detect_emotion_deepface(self, frame):
        """ตรวจจับอารมณ์ด้วย DeepFace แบบ real-time"""
        try:
            if not DEEPFACE_AVAILABLE:
                return self.detect_faces_simple(frame)
            
            # ตรวจสอบแคช
            frame_hash = hash(frame.tobytes())
            if frame_hash in self.emotion_cache:
                return self.emotion_cache[frame_hash]
            
            # ตรวจจับทุก 50ms (20 FPS) สำหรับ real-time
            current_time = time.time()
            if current_time - self.last_emotion_time < DETECTION_INTERVAL:
                return None
            
            self.last_emotion_time = current_time
            
            if self.camera_method == "picamera2":
                if len(frame.shape) == 3 and frame.shape[2] == 3:
                    frame_bgr = cv2.cvtColor(frame, cv2.COLOR_RGB2BGR)
                else:
                    frame_bgr = frame
            else:
                frame_bgr = frame
                
            result = DeepFace.analyze(
                frame_bgr, 
                actions=['emotion'], 
                enforce_detection=False,
                silent=True
            )
            
            if isinstance(result, list) and len(result) > 0:
                emotion_map = {
                    'angry': 'Angry',
                    'disgust': 'Disgust',
                    'fear': 'Fear',
                    'happy': 'Happy',
                    'sad': 'Sad',
                    'surprise': 'Surprise',
                    'neutral': 'Neutral'
                }
                
                emotion = result[0]['dominant_emotion']
                confidence = result[0]['emotion'][emotion]
                
                # วาดกรอบใบหน้าจาก DeepFace
                if 'region' in result[0]:
                    region = result[0]['region']
                    x, y, w, h = region['x'], region['y'], region['w'], region['h']
                    cv2.rectangle(frame_bgr, (x, y), (x+w, y+h), (0, 255, 0), 3)
                    cv2.putText(frame_bgr, f"AI: {emotion}", (x, y-15), 
                               cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
                
                # ใช้การประเมินใหม่
                assessed_emotion, satisfaction_text, satisfaction_level = self.get_emotion_assessment(confidence)
                
                result = (
                    assessed_emotion,
                    confidence,
                    satisfaction_level,
                    satisfaction_text
                )
                
                # เก็บผลลัพธ์ในแคช
                self.emotion_cache[frame_hash] = result
                if len(self.emotion_cache) > EMOTION_CACHE_SIZE:
                    self.emotion_cache.pop(next(iter(self.emotion_cache)))
                
                # บันทึกข้อมูล (ทุก 5 วินาที)
                self.save_to_excel(*result)
                
                return result
            else:
                # ถ้า DeepFace ไม่เจอหน้า ให้ใช้ simple detection
                return self.detect_faces_simple(frame)
                
        except Exception as e:
            print(f"DeepFace error: {e}")
            return self.detect_faces_simple(frame)
    
    def get_current_emotion(self):
        """รับอารมณ์ปัจจุบัน (AI หรือ Manual)"""
        current_time = time.time()
        
        # ตรวจสอบ manual emotion (ใช้ได้ 5 วินาที)
        if (self.manual_emotion and 
            current_time - self.last_manual_time < 5.0):
            assessed_emotion, satisfaction_text, satisfaction_level = self.get_emotion_assessment(self.manual_confidence)
            return (
                f"[Manual] {assessed_emotion}",
                self.manual_confidence,
                satisfaction_level,
                satisfaction_text
            )
        
        return None
    
    def detect_faces_simple(self, frame):
        """ตรวจจับใบหน้าแบบง่าย พร้อมจัดการสีและกรอบแสดงผล"""
        try:
            if self.face_cascade is None:
                return "Face Detected", 50.0, 3, "***"
            
            # ปรับปรุงเฟรมเพื่อการตรวจจับที่ดีขึ้น
            enhanced_frame = self.enhance_frame_for_detection(frame)
            
            # แปลงเป็น grayscale สำหรับ face detection
            if len(enhanced_frame.shape) == 3:
                gray = cv2.cvtColor(enhanced_frame, cv2.COLOR_BGR2GRAY)
            else:
                gray = enhanced_frame
            
            # ปรับปรุงพารามิเตอร์การตรวจจับใบหน้า
            faces = self.face_cascade.detectMultiScale(
                gray, 
                scaleFactor=1.05,  # ลดลงเพื่อตรวจจับได้ดีขึ้น
                minNeighbors=2,    # ลดลงเพื่อตรวจจับได้ง่ายขึ้น
                minSize=(30, 30),  # เพิ่มขนาดขั้นต่ำเล็กน้อย
                maxSize=(300, 300), # เพิ่มขนาดสูงสุด
                flags=cv2.CASCADE_SCALE_IMAGE
            )
            
            # วาดกรอบรอบใบหน้าที่ตรวจพบ
            if len(faces) > 0:
                # หาใบหน้าที่ใหญ่ที่สุด (น่าจะเป็นใบหน้าหลัก)
                largest_face = max(faces, key=lambda face: face[2] * face[3])
                x, y, w, h = largest_face
                
                # วาดกรอบใบหน้าหลัก (สีเขียวเข้ม)
                cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 255, 0), 3)
                
                # เพิ่มป้ายกำกับใบหน้า
                cv2.putText(frame, "FACE DETECTED", (x, y-15), 
                           cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
                
                # วาดกรอบใบหน้าอื่นๆ (สีเขียวอ่อน)
                for (fx, fy, fw, fh) in faces:
                    if (fx, fy, fw, fh) != (x, y, w, h):  # ไม่ใช่ใบหน้าหลัก
                        cv2.rectangle(frame, (fx, fy), (fx+fw, fy+fh), (0, 200, 0), 1)
                        cv2.putText(frame, "Face", (fx, fy-5), 
                                   cv2.FONT_HERSHEY_SIMPLEX, 0.4, (0, 200, 0), 1)
                
                # แสดงจำนวนใบหน้าที่ตรวจจับได้
                cv2.putText(frame, f"Faces: {len(faces)}", (10, 30), 
                           cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)
                
                # คืนค่าการประเมินอารมณ์สำหรับใบหน้าที่ตรวจพบ
                assessed_emotion, satisfaction_text, satisfaction_level = self.get_emotion_assessment(65.0)
                result = (assessed_emotion, 65.0, satisfaction_level, satisfaction_text)
                
                # บันทึกข้อมูล (ทุก 5 วินาที)
                self.save_to_excel(*result)
                
                return result
            else:
                # แสดงข้อความ "No Face Detected"
                height, width = frame.shape[:2]
                cv2.putText(frame, "NO FACE DETECTED", (width//2 - 150, height//2), 
                           cv2.FONT_HERSHEY_SIMPLEX, 1.0, (0, 0, 255), 2)
                cv2.putText(frame, "Please position your face in front of camera", 
                           (width//2 - 200, height//2 + 40), 
                           cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 255), 1)
                return "No Face", 0.0, 0, ""
                
        except Exception as e:
            print(f"Face detection error: {e}")
            return "Error", 0.0, 0, ""
    
    def add_overlay_info(self, frame, emotion, confidence, satisfaction_level, satisfaction_text):
        """เพิ่มข้อมูลลงบนเฟรม พร้อมจัดการสี"""
        if frame is None:
            return frame
        
        height, width = frame.shape[:2]
        
        # วาดกรอบข้อมูลพื้นหลังโปร่งใส (ปรับขนาดตามหน้าจอ)
        overlay_height = max(200, int(height * 0.3))  # 30% ของความสูงหน้าจอ
        overlay = frame.copy()
        cv2.rectangle(overlay, (10, 10), (width-10, overlay_height), (0, 0, 0), -1)
        cv2.addWeighted(overlay, 0.7, frame, 0.3, 0, frame)
        
        # กรอบขอบ
        cv2.rectangle(frame, (10, 10), (width-10, overlay_height), (255, 255, 255), 2)
        
        # คำนวณขนาดข้อความตามขนาดหน้าจอ
        font_scale = max(0.8, min(2.0, width / 800))  # ปรับขนาดตามความกว้างหน้าจอ
        font_thickness = max(1, int(width / 400))  # ปรับความหนาตามขนาดหน้าจอ
        
        # ข้อความอารมณ์ (สีเขียวสดใส)
        emotion_text = f"Emotion: {emotion}"
        cv2.putText(frame, emotion_text, (20, 40), 
                   cv2.FONT_HERSHEY_SIMPLEX, font_scale, (0, 255, 0), font_thickness)
        
        # ความมั่นใจ (สีเหลือง)
        if isinstance(confidence, (int, float)) and confidence > 0:
            conf_text = f"Confidence: {confidence:.2f}%"
            cv2.putText(frame, conf_text, (20, 70), 
                       cv2.FONT_HERSHEY_SIMPLEX, font_scale * 0.8, (0, 255, 255), font_thickness)
        
        # ระดับความพึงพอใจ (สีฟ้า)
        satisfaction_level_text = f"Satisfaction: {satisfaction_level}/5"
        cv2.putText(frame, satisfaction_level_text, (20, 100), 
                   cv2.FONT_HERSHEY_SIMPLEX, font_scale * 0.8, (255, 255, 0), font_thickness)
        
        # ดาวแสดงความพึงพอใจ (สีเหลือง) - ใช้ * แทน ★
        stars_text = f"Score: {satisfaction_text}"
        cv2.putText(frame, stars_text, (20, 130), 
                   cv2.FONT_HERSHEY_SIMPLEX, font_scale * 0.8, (0, 255, 255), font_thickness)
        
        # วิธีการกล้อง (สีขาว)
        camera_text = f"Camera: {self.camera_method}"
        cv2.putText(frame, camera_text, (20, 160), 
                   cv2.FONT_HERSHEY_SIMPLEX, font_scale * 0.6, (255, 255, 255), font_thickness)
        
        # โหมดสี (สีม่วง)
        color_text = f"Mode: {self.color_mode}"
        cv2.putText(frame, color_text, (20, 190), 
                   cv2.FONT_HERSHEY_SIMPLEX, font_scale * 0.6, (255, 0, 255), font_thickness)
        
        # แสดง manual controls
        manual_text = "Manual: 1=Happy 2=Neutral 3=Sad 4=Angry"
        cv2.putText(frame, manual_text, (20, 220), 
                   cv2.FONT_HERSHEY_SIMPLEX, font_scale * 0.5, (255, 255, 255), font_thickness)
        
        # แสดงสถานะ manual mode
        if self.manual_emotion and time.time() - self.last_manual_time < 5.0:
            manual_status = f"Manual Active: {self.manual_emotion}"
            cv2.putText(frame, manual_status, (20, 250), 
                       cv2.FONT_HERSHEY_SIMPLEX, font_scale * 0.6, (0, 255, 255), font_thickness)
        
        # แสดงสถานะการบันทึกข้อมูล
        current_time = time.time()
        time_since_save = current_time - self.last_data_save_time
        save_status = f"Next save in: {DATA_SAVE_INTERVAL - time_since_save:.1f}s"
        cv2.putText(frame, save_status, (20, 280), 
                   cv2.FONT_HERSHEY_SIMPLEX, font_scale * 0.5, (255, 255, 255), font_thickness)
        
        # เวลา (สีฟ้า) - ปรับตำแหน่งให้เหมาะสมกับหน้าจอใหญ่
        time_text = datetime.now().strftime("%H:%M:%S")
        time_x = max(width - 200, width - int(width * 0.3))
        cv2.putText(frame, time_text, (time_x, 40), 
                   cv2.FONT_HERSHEY_SIMPLEX, font_scale * 0.8, (255, 255, 0), font_thickness)
        
        return frame
    
    def save_emotion_data(self, emotion, confidence):
        """บันทึกข้อมูลอารมณ์"""
        timestamp = datetime.now().isoformat()
        data = {
            'timestamp': timestamp,
            'emotion': emotion,
            'confidence': confidence,
            'camera_method': self.camera_method
        }
        self.emotion_history.append(data)
        
        if len(self.emotion_history) > 100:
            self.emotion_history.pop(0)
    
    def run(self):
        """เริ่มการทำงานหลัก"""
        print("🎭 เริ่มการตรวจจับอารมณ์ด้วยกล้อง")
        print("=" * 60)
        
        if not self.setup_camera():
            return
        
        print(f"📷 เริ่มต้นกล้องสำเร็จ: {self.camera_method}")
        print("🎯 เริ่มการตรวจจับอารมณ์แบบ Real-time...")
        print("💾 ข้อมูลจะถูกบันทึกทุก 5 วินาที")
        print(f"📁 Excel file: {os.path.abspath(self.excel_file)}")
        print("📋 คำสั่งควบคุม:")
        print("   - กด 'q' เพื่อออก")
        print("   - กด 's' หรือ SPACE เพื่อบันทึกรูปภาพ")
        print("   - กด 'i' เพื่อดูข้อมูลกล้อง")
        print("   - กด 'c' เพื่อสลับโหมดสี")
        print("   - กด 'f' เพื่อสลับโหมดเต็มจอ")
        print("   - กด 'b'/'v' เพื่อปรับความสว่าง (+/-)")
        print("   - กด 'n'/'m' เพื่อปรับคอนทราสต์ (+/-)")
        print("")
        print("🎯 Manual Emotion Input:")
        print("   - Press '1' = Happy (80-100%)")
        print("   - Press '2' = Neutral (50-70%)")
        print("   - Press '3' = Sad (30-40%)")
        print("   - Press '4' = Angry (0-20%)")
        print("=" * 60)
        
        self.is_running = True
        frame_count = 0
        fps_start_time = time.time()
        last_fps_update = time.time()
        fps = 0
        
        try:
            while True:
                frame = self.get_frame()
                
                if frame is None:
                    print("❌ Error: Can't receive frame")
                    time.sleep(0.1)
                    continue
                
                frame_count += 1
                
                # ไม่ข้ามเฟรม (real-time detection)
                # if frame_count % FRAME_SKIP != 0:
                #     continue
                
                # ตรวจสอบ manual emotion ก่อน
                manual_result = self.get_current_emotion()
                if manual_result:
                    emotion, confidence, satisfaction_level, satisfaction_text = manual_result
                else:
                    # ตรวจจับอารมณ์ด้วย AI
                    result = self.detect_emotion_deepface(frame)
                    if result:
                        emotion, confidence, satisfaction_level, satisfaction_text = result
                    else:
                        emotion, confidence, satisfaction_level, satisfaction_text = "No Face", 0.0, 0, ""
                
                display_frame = self.add_overlay_info(
                    frame, emotion, confidence, satisfaction_level, satisfaction_text
                )
                
                if display_frame is not None:
                    # แสดงหน้าต่างเต็มจอหรือปกติ
                    cv2.namedWindow('Emotion Detection', cv2.WINDOW_NORMAL)
                    if self.fullscreen_mode:
                        cv2.setWindowProperty('Emotion Detection', cv2.WND_PROP_FULLSCREEN, cv2.WINDOW_FULLSCREEN)
                    else:
                        cv2.setWindowProperty('Emotion Detection', cv2.WND_PROP_FULLSCREEN, cv2.WINDOW_NORMAL)
                    cv2.imshow('Emotion Detection', display_frame)
                
                # อัพเดท FPS ทุก 1 วินาที
                current_time = time.time()
                if current_time - last_fps_update >= 1.0:
                    fps = frame_count / (current_time - fps_start_time)
                    frame_count = 0
                    fps_start_time = current_time
                    last_fps_update = current_time
                    print(f"📊 FPS: {fps:.1f}")
                
                # จัดการ key input
                key = cv2.waitKey(1) & 0xFF
                if key == ord('q'):
                    break
                elif key == ord('s') or key == 32:
                    filename = f"emotion_screenshot_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg"
                    cv2.imwrite(filename, display_frame)
                    print(f"📸 Screenshot saved: {filename}")
                elif key == ord('i'):
                    self.show_camera_info()
                elif key == ord('c'):
                    self.toggle_color_mode()
                elif key == ord('f'):
                    self.toggle_fullscreen()
                elif key == ord('b'):
                    self.adjust_brightness(0.1)
                elif key == ord('v'):
                    self.adjust_brightness(-0.1)
                elif key == ord('n'):
                    self.adjust_contrast(0.1)
                elif key == ord('m'):
                    self.adjust_contrast(-0.1)
                # Manual emotion input
                elif key == ord('1'):
                    self.set_manual_emotion("Happy", 90.0)
                elif key == ord('2'):
                    self.set_manual_emotion("Neutral", 60.0)
                elif key == ord('3'):
                    self.set_manual_emotion("Sad", 35.0)
                elif key == ord('4'):
                    self.set_manual_emotion("Angry", 10.0)
                
        except KeyboardInterrupt:
            print("\n⚠️ Interrupted by user")
        
        finally:
            self.cleanup()
    
    def toggle_color_mode(self):
        """สลับโหมดสี"""
        if self.color_mode == "color":
            self.color_mode = "grayscale"
            print("🎨 Switched to grayscale mode")
        else:
            self.color_mode = "color"
            print("🎨 Switched to color mode")
    
    def toggle_fullscreen(self):
        """สลับโหมดเต็มจอ"""
        self.fullscreen_mode = not self.fullscreen_mode
        if self.fullscreen_mode:
            print("🖥️ Switched to fullscreen mode")
        else:
            print("🖥️ Switched to window mode")
    
    def adjust_brightness(self, delta):
        """ปรับความสว่าง"""
        self.brightness = max(-1.0, min(1.0, self.brightness + delta))
        print(f"💡 Brightness: {self.brightness:.1f}")
        
        if self.picam2:
            try:
                self.picam2.set_controls({"Brightness": self.brightness})
            except:
                pass
    
    def adjust_contrast(self, delta):
        """ปรับคอนทราสต์"""
        self.contrast = max(0.0, min(2.0, self.contrast + delta))
        print(f"🔆 Contrast: {self.contrast:.1f}")
        
        if self.picam2:
            try:
                self.picam2.set_controls({"Contrast": self.contrast})
            except:
                pass
    
    def show_camera_info(self):
        """แสดงข้อมูลกล้อง"""
        print("\n📷 Camera Information:")
        print(f"   Method: {self.camera_method}")
        print(f"   Emotion history: {len(self.emotion_history)} records")
        print(f"   DeepFace available: {DEEPFACE_AVAILABLE}")
        print(f"   PiCamera2 available: {PICAMERA2_AVAILABLE}")
        
        if self.cap:
            width = self.cap.get(cv2.CAP_PROP_FRAME_WIDTH)
            height = self.cap.get(cv2.CAP_PROP_FRAME_HEIGHT)
            fps = self.cap.get(cv2.CAP_PROP_FPS)
            print(f"   Resolution: {int(width)}x{int(height)}")
            print(f"   FPS Setting: {fps}")
    
    def cleanup(self):
        """ทำความสะอาดและบันทึกข้อมูลสุดท้าย"""
        print("🧹 Cleaning up...")
        
        self.is_running = False
        
        if self.picam2:
            try:
                self.picam2.stop()
            except:
                pass
        
        if self.cap:
            self.cap.release()
        
        cv2.destroyAllWindows()
        
        # บันทึกข้อมูลที่เหลือในคิว
        try:
            while not self.data_queue.empty():
                data = self.data_queue.get_nowait()
                self._save_to_excel_internal(*data)
            self._save_excel_file()
        except:
            pass
        
        if self.emotion_history:
            print(f"📊 Total emotions detected: {len(self.emotion_history)}")
            
            emotions = [record['emotion'] for record in self.emotion_history]
            unique_emotions = set(emotions)
            print("📈 Emotion statistics:")
            for emotion in unique_emotions:
                count = emotions.count(emotion)
                percentage = (count / len(emotions)) * 100
                print(f"   {emotion}: {count} times ({percentage:.1f}%)")
        
        print("✅ Cleanup completed")
    
    def test_face_detection(self):
        """ทดสอบการตรวจจับใบหน้า"""
        print("🧪 Testing face detection...")
        
        if not self.setup_camera():
            print("❌ Camera setup failed")
            return False
        
        print("📷 Camera ready, testing face detection...")
        print("Press 'q' to quit, 's' to save screenshot")
        
        frame_count = 0
        face_detected_count = 0
        
        try:
            while True:
                frame = self.get_frame()
                if frame is None:
                    continue
                
                frame_count += 1
                
                # ทดสอบการตรวจจับใบหน้า
                result = self.detect_faces_simple(frame)
                if result and result[0] != "No Face":
                    face_detected_count += 1
                
                # แสดงผล
                cv2.imshow('Face Detection Test', frame)
                
                # แสดงสถิติ
                if frame_count % 30 == 0:  # ทุก 30 เฟรม
                    detection_rate = (face_detected_count / frame_count) * 100
                    print(f"📊 Detection rate: {detection_rate:.1f}% ({face_detected_count}/{frame_count})")
                
                key = cv2.waitKey(1) & 0xFF
                if key == ord('q'):
                    break
                elif key == ord('s'):
                    filename = f"face_test_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg"
                    cv2.imwrite(filename, frame)
                    print(f"📸 Screenshot saved: {filename}")
        
        except KeyboardInterrupt:
            print("\n⚠️ Test interrupted")
        
        finally:
            cv2.destroyAllWindows()
            if self.cap:
                self.cap.release()
            
            # สรุปผลการทดสอบ
            if frame_count > 0:
                detection_rate = (face_detected_count / frame_count) * 100
                print(f"\n📊 Final Detection Rate: {detection_rate:.1f}%")
                print(f"📈 Frames processed: {frame_count}")
                print(f"👤 Faces detected: {face_detected_count}")
                
                if detection_rate > 80:
                    print("✅ Face detection working well!")
                elif detection_rate > 50:
                    print("⚠️ Face detection working but could be better")
                else:
                    print("❌ Face detection needs improvement")
                    print("💡 Try:")
                    print("   - Better lighting")
                    print("   - Face directly in front of camera")
                    print("   - Remove glasses or hat")
                    print("   - Check camera focus")
            
            return detection_rate > 50 if frame_count > 0 else False

def main():
    print("🎭 ระบบตรวจจับอารมณ์ด้วยกล้อง")
    print("🎨 เวอร์ชัน 2.4 - Real-time Detection + Data Save Every 5s")
    print("=" * 70)
    
    # เลือกโหมดการทำงาน
    print("\n🔧 เลือกโหมดการทำงาน:")
    print("1. ระบบตรวจจับอารมณ์ปกติ")
    print("2. ทดสอบการตรวจจับใบหน้า")
    
    while True:
        try:
            mode_choice = input("เลือกโหมด (1-2): ").strip()
            if mode_choice in ["1", "2"]:
                break
            else:
                print("❌ ตัวเลือกไม่ถูกต้อง กรุณาเลือก 1 หรือ 2")
        except:
            print("❌ การป้อนข้อมูลไม่ถูกต้อง กรุณาลองใหม่")
    
    # เลือกประเภทกล้อง
    print("\n🔧 เลือกประเภทกล้อง:")
    print("1. กล้องเว็บแคม (โน๊ตบุ๊ค)")
    print("2. กล้อง Raspberry Pi")
    
    while True:
        try:
            choice = input("เลือกประเภทกล้อง (1-2): ").strip()
            if choice == "1":
                camera_type = "laptop"
                break
            elif choice == "2":
                camera_type = "pi"
                break
            else:
                print("❌ ตัวเลือกไม่ถูกต้อง กรุณาเลือก 1 หรือ 2")
        except:
            print("❌ การป้อนข้อมูลไม่ถูกต้อง กรุณาลองใหม่")
    
    # เลือกโหมดสี
    print("\n🔧 เลือกโหมดการแสดงผล:")
    print("1. โหมดสี (ค่าเริ่มต้น)")
    print("2. โหมดขาวดำ")
    
    try:
        choice = input("เลือกโหมด (1-2, กด Enter สำหรับค่าเริ่มต้น): ").strip()
        color_mode = "grayscale" if choice == "2" else "color"
    except:
        color_mode = "color"
    
    detector = RaspberryPi4CameraDetector()
    detector.camera_type = camera_type
    detector.color_mode = color_mode
    detector.check_camera_hardware()
    
    if mode_choice == "1":
        detector.run()
    else:
        detector.test_face_detection()

if __name__ == "__main__":
    main()
    